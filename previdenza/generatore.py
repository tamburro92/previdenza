"""
Generazione file Excel con risultati calcolo contributi
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class GeneratoreExcel:
    """Classe per generare il file Excel con i risultati"""

    def __init__(self, risultati_calcolo):
        self.risultati = risultati_calcolo
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Contributi Previdenziali"

    def genera(self, output_path):
        """Genera il file Excel"""
        self._applica_stili()
        self._crea_headers()
        self._popola_dati()
        self._aggiungi_totali()
        self._imposta_larghezza_colonne()
        self.wb.save(output_path)

    def _applica_stili(self):
        """Definisce gli stili"""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill("solid", fgColor="4472C4")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center = Alignment(horizontal='center')

    def _crea_headers(self):
        """Crea le intestazioni"""
        headers = ["Anno", "Giorni REALI", "", "Anno", "Giorni TEORICI", "Mesi", "Anni e Mesi Cumulativi"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            if header:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center
                cell.border = self.border

    def _popola_dati(self):
        """Popola i dati per ogni anno"""
        anno_min = self.risultati["anno_min"]
        anno_max = self.risultati["anno_max"]

        if not anno_min or not anno_max:
            return

        mesi_cumulativi = 0

        for i, anno in enumerate(range(anno_min, anno_max + 1), 2):
            reale = self.risultati["reale"].get(anno, 0)
            teorico = self.risultati["teorico"].get(anno, 0)
            mesi = self.risultati["mesi"].get(anno, 0)

            # Colonne REALI (A, B)
            self.ws.cell(row=i, column=1, value=anno).border = self.border
            self.ws.cell(row=i, column=1).alignment = self.center
            self.ws.cell(row=i, column=2, value=reale).border = self.border
            self.ws.cell(row=i, column=2).alignment = self.center

            # Colonne TEORICI (D, E)
            self.ws.cell(row=i, column=4, value=anno).border = self.border
            self.ws.cell(row=i, column=4).alignment = self.center
            self.ws.cell(row=i, column=5, value=teorico).border = self.border
            self.ws.cell(row=i, column=5).alignment = self.center

            # Mesi (F)
            self.ws.cell(row=i, column=6, value=mesi).border = self.border
            self.ws.cell(row=i, column=6).alignment = self.center

            # Anni e Mesi Cumulativi (G)
            mesi_cumulativi += mesi
            anni_cum = mesi_cumulativi // 12
            mesi_cum = mesi_cumulativi % 12
            formato_cumulativo = f"{anni_cum}a {mesi_cum}m"
            self.ws.cell(row=i, column=7, value=formato_cumulativo).border = self.border
            self.ws.cell(row=i, column=7).alignment = self.center

    def _aggiungi_totali(self):
        """Aggiunge la riga dei totali"""
        anno_min = self.risultati["anno_min"]
        anno_max = self.risultati["anno_max"]

        if not anno_min or not anno_max:
            return

        num_anni = anno_max - anno_min + 1
        last_row = num_anni + 2

        # Totale REALE
        self.ws.cell(row=last_row, column=1, value="TOTALE").font = Font(bold=True)
        self.ws.cell(row=last_row, column=1).border = self.border
        self.ws.cell(row=last_row, column=2, value=f"=SUM(B2:B{last_row-1})").font = Font(bold=True)
        self.ws.cell(row=last_row, column=2).border = self.border
        self.ws.cell(row=last_row, column=2).alignment = self.center

        # Totale TEORICO
        self.ws.cell(row=last_row, column=4, value="TOTALE").font = Font(bold=True)
        self.ws.cell(row=last_row, column=4).border = self.border
        self.ws.cell(row=last_row, column=5, value=f"=SUM(E2:E{last_row-1})").font = Font(bold=True)
        self.ws.cell(row=last_row, column=5).border = self.border
        self.ws.cell(row=last_row, column=5).alignment = self.center

        # Totale Mesi
        self.ws.cell(row=last_row, column=6, value=f"=SUM(F2:F{last_row-1})").font = Font(bold=True)
        self.ws.cell(row=last_row, column=6).border = self.border
        self.ws.cell(row=last_row, column=6).alignment = self.center

        # Label finale per Anni e Mesi Cumulativi (usa obiettivo dal calcolo)
        obiettivo_label = self.risultati.get("obiettivo_label", "42a 10m")
        self.ws.cell(row=last_row, column=7, value=obiettivo_label).font = Font(bold=True)
        self.ws.cell(row=last_row, column=7).border = self.border
        self.ws.cell(row=last_row, column=7).alignment = self.center

    def _imposta_larghezza_colonne(self):
        """Imposta la larghezza delle colonne"""
        self.ws.column_dimensions['A'].width = 12
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 5
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 18
        self.ws.column_dimensions['F'].width = 10
        self.ws.column_dimensions['G'].width = 24
