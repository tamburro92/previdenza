"""
===============================================================================
CALCOLO CONTRIBUTI PREVIDENZIALI INPS
===============================================================================
Script per estrarre dati da PDF Estratto Conto Previdenziale INPS,
calcolare i contributi REALI e TEORICI, e generare Excel e JSON.

Uso:
    python contributi_inps.py <percorso_pdf> [--output-dir <cartella>]

Esempio:
    python contributi_inps.py certificazione.pdf
    python contributi_inps.py certificazione.pdf --output-dir ./risultati

Output:
    - contributi_estratti.json  (dati grezzi estratti dal PDF)
    - contributi_previdenziali.xlsx (calcolo REALE e TEORICO per anno)

Requisiti:
    pip install pdfplumber openpyxl
===============================================================================
"""

import pdfplumber
import json
import re
import sys
import os
import argparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# UTILITA' CODICE FISCALE
# =============================================================================

def decodifica_sesso_da_cf(codice_fiscale):
    """
    Decodifica il sesso dal codice fiscale italiano.
    Il giorno di nascita e' nelle posizioni 9-10 (indice 9-11).
    Se il giorno > 40, e' donna (si aggiunge 40 al giorno reale).

    Returns: 'F' per femmina, 'M' per maschio, None se non decodificabile
    """
    if not codice_fiscale or len(codice_fiscale) < 11:
        return None

    try:
        giorno = int(codice_fiscale[9:11])
        if giorno > 40:
            return 'F'  # Donna
        else:
            return 'M'  # Uomo
    except (ValueError, IndexError):
        return None


# =============================================================================
# ESTRAZIONE PDF
# =============================================================================

class EstrattorePDF:
    """Classe per estrarre i dati contributivi da PDF INPS"""
    
    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.dati = {
            "regime_generale": [],
            "spettacolo": [],
            "metadata": {
                "file": pdf_path,
                "codice_fiscale": None,
                "cognome": None,
                "nome": None
            }
        }
    
    def estrai(self):
        """Estrae tutti i dati dal PDF"""
        with pdfplumber.open(self.pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                self._estrai_metadata(page, page_num)
                self._estrai_tabelle(page)
        
        return self.dati
    
    def _estrai_metadata(self, page, page_num):
        """Estrae metadata dalla prima pagina"""
        if page_num == 0:
            text = page.extract_text()

            # Estrai codice fiscale
            cf_match = re.search(r'([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])', text)
            if cf_match:
                self.dati["metadata"]["codice_fiscale"] = cf_match.group(1)

            # Estrai cognome e nome da "Estratto conto di COGNOME NOME CODICEFISCALE"
            nome_match = re.search(r'Estratto\s+conto\s+di\s+([A-Z][A-Z\s]+?)\s+([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])', text)
            if nome_match:
                nome_completo = nome_match.group(1).strip()
                parti = nome_completo.split()
                if len(parti) >= 2:
                    self.dati["metadata"]["cognome"] = parti[0]
                    self.dati["metadata"]["nome"] = ' '.join(parti[1:])
    
    def _estrai_tabelle(self, page):
        """Estrae le tabelle dalla pagina"""
        tables = page.extract_tables()
        
        for table in tables:
            if not table or len(table) < 3:
                continue
            
            header = table[0] if table[0] else []
            header_str = ' '.join([str(h) for h in header if h])
            data_rows = table[2:] if len(table) > 2 else []
            
            for row in data_rows:
                self._processa_riga(row, header_str)
    
    def _processa_riga(self, row, header_str):
        """Processa una singola riga della tabella"""
        if not row or not row[0]:
            return
        
        dal = row[0]
        al = row[1]
        tipo = row[2]
        
        # Verifica data valida
        if not re.match(r'\d{2}/\d{2}/\d{4}', str(dal)):
            return
        
        # Regime Generale (settimane)
        if 'sett.' in str(row[3]) or row[3] == 'sett.':
            self._processa_regime_generale(row, dal, al, tipo)
        
        # Spettacolo (giorni)
        elif 'Giorni' in header_str or 'P.A.L.S.' in str(tipo) or 'Malattia' in str(tipo):
            self._processa_spettacolo(row, dal, al, tipo)
    
    def _processa_regime_generale(self, row, dal, al, tipo):
        """Processa record Regime Generale"""
        settimane = None
        for val in row[3:]:
            if val and re.match(r'^\d+$', str(val)):
                settimane = int(val)
                break
        
        if settimane:
            record = {
                "dal": dal,
                "al": al,
                "tipo": tipo,
                "settimane": settimane,
                "unita": "settimane"
            }
            self._aggiungi_retribuzione(record, row)
            self.dati["regime_generale"].append(record)
    
    def _processa_spettacolo(self, row, dal, al, tipo):
        """Processa record Lavoratori Spettacolo"""
        giorni = None
        gruppo = None
        codice_qualifica = None
        retribuzione = None
        
        # Giorni in posizione 3
        if row[3] and str(row[3]).strip():
            try:
                giorni = int(row[3])
            except:
                pass
        
        # Retribuzione in posizione 4
        if len(row) > 4 and row[4]:
            retribuzione = row[4]
        
        # Gruppo in posizione 6
        if len(row) > 6 and row[6]:
            try:
                gruppo = int(row[6])
            except:
                pass
        
        # Codice qualifica in posizione 7
        if len(row) > 7 and row[7]:
            codice_qualifica = str(row[7]).replace('\n', '')
        
        record = {
            "dal": dal,
            "al": al,
            "tipo": tipo,
            "giorni": giorni,
            "unita": "giorni"
        }
        
        if gruppo:
            record["gruppo"] = gruppo
        if codice_qualifica:
            record["codice_qualifica"] = codice_qualifica
        if retribuzione:
            record["retribuzione"] = retribuzione
        
        self.dati["spettacolo"].append(record)
    
    def _aggiungi_retribuzione(self, record, row):
        """Aggiunge la retribuzione al record se presente"""
        for val in row:
            if val and re.match(r'[\d.,]+$', str(val).replace('.', '').replace(',', '')):
                try:
                    ret = float(str(val).replace('.', '').replace(',', '.'))
                    if ret > 100:
                        record["retribuzione"] = val
                        break
                except:
                    pass


# =============================================================================
# CALCOLO CONTRIBUTI
# =============================================================================

class CalcolatoreContributi:
    """Classe per calcolare i contributi REALI e TEORICI"""

    # Obiettivi contributivi per sesso
    OBIETTIVO_DONNA = 41 * 12 + 10  # 41 anni e 10 mesi = 502 mesi
    OBIETTIVO_UOMO = 42 * 12 + 10   # 42 anni e 10 mesi = 514 mesi

    def __init__(self, dati_estratti, sesso=None, tempo_indeterminato_da=None):
        self.dati = dati_estratti
        self.sesso = sesso  # 'M' o 'F'
        self.tempo_indeterminato_da = tempo_indeterminato_da  # None, "sempre", o "DD/MM/YYYY"
        self.reale_per_anno = defaultdict(int)
        self.teorico_per_anno = defaultdict(int)
        self.mesi_per_anno = defaultdict(int)  # Mesi teorici per anno
        self.ultimo_regime = None  # Per estensione anni futuri
        self.ultimo_gruppo = None  # Gruppo spettacolo se applicabile
        self.anno_min = None
        self.anno_max = None

        # Determina obiettivo in base al sesso
        if sesso == 'F':
            self.obiettivo_mesi = self.OBIETTIVO_DONNA
            self.obiettivo_label = "41a 10m"
        else:
            self.obiettivo_mesi = self.OBIETTIVO_UOMO
            self.obiettivo_label = "42a 10m"

    def _is_tempo_indeterminato(self, anno, mese):
        """Verifica se in un dato anno/mese il contratto e' a tempo indeterminato"""
        if self.tempo_indeterminato_da is None:
            return False  # Sempre tempo determinato
        if self.tempo_indeterminato_da == "sempre":
            return True  # Sempre tempo indeterminato
        # Altrimenti confronta con la data
        g, m, a = map(int, self.tempo_indeterminato_da.split('/'))
        if anno > a:
            return True
        if anno == a and mese >= m:
            return True
        return False

    def calcola(self):
        """Esegue tutti i calcoli"""
        self._calcola_regime_generale()
        self._calcola_spettacolo()
        self._determina_range_anni()
        self._estendi_a_obiettivo()

        return {
            "reale": dict(self.reale_per_anno),
            "teorico": dict(self.teorico_per_anno),
            "mesi": dict(self.mesi_per_anno),
            "anno_min": self.anno_min,
            "anno_max": self.anno_max,
            "sesso": self.sesso,
            "obiettivo_mesi": self.obiettivo_mesi,
            "obiettivo_label": self.obiettivo_label
        }
    
    def _parse_data(self, data_str):
        """Converte stringa data in (anno, mese, giorno)"""
        giorno, mese, anno = map(int, data_str.split('/'))
        return anno, mese, giorno
    
    def _conta_mesi(self, dal, al):
        """Conta i mesi tra due date (arrotondato per eccesso)"""
        y1, m1, _ = self._parse_data(dal)
        y2, m2, _ = self._parse_data(al)
        return (y2 - y1) * 12 + (m2 - m1) + 1
    
    def _calcola_regime_generale(self):
        """Calcola contributi Regime Generale"""
        for record in self.dati["regime_generale"]:
            anno = self._parse_data(record["dal"])[0]
            settimane = record["settimane"]

            # REALE: settimane * 6
            giorni_reali = settimane * 6
            self.reale_per_anno[anno] += giorni_reali

            # TEORICO: mesi * 26
            mesi = self._conta_mesi(record["dal"], record["al"])
            giorni_teorici = mesi * 26
            self.teorico_per_anno[anno] += giorni_teorici
            self.mesi_per_anno[anno] += mesi

            self.ultimo_regime = "generale"
    
    def _calcola_spettacolo(self):
        """Calcola contributi Lavoratori Spettacolo"""
        # Raccogli periodi per anno: con gruppo e senza gruppo separatamente
        periodi_con_gruppo = defaultdict(list)  # anno -> [(mese_inizio, mese_fine, gruppo)]
        periodi_senza_gruppo = defaultdict(list)  # anno -> [(mese_inizio, mese_fine, giorni)]
        giorni_senza_gruppo_per_anno = defaultdict(int)  # anno -> giorni totali (per record senza gruppo)

        for record in self.dati["spettacolo"]:
            anno_inizio, mese_inizio, _ = self._parse_data(record["dal"])
            anno_fine, mese_fine, _ = self._parse_data(record["al"])
            giorni = record.get("giorni")
            gruppo = record.get("gruppo")

            # REALE: somma tutti i giorni (inclusi malattia, maternita', ecc.)
            if giorni:
                self.reale_per_anno[anno_inizio] += giorni

            if gruppo:
                # Record CON gruppo: calcola in base alle regole spettacolo
                if anno_inizio == anno_fine:
                    periodi_con_gruppo[anno_inizio].append((mese_inizio, mese_fine, gruppo))
                else:
                    # Periodo che attraversa anni: spezza per anno
                    periodi_con_gruppo[anno_inizio].append((mese_inizio, 12, gruppo))
                    for anno in range(anno_inizio + 1, anno_fine):
                        periodi_con_gruppo[anno].append((1, 12, gruppo))
                    periodi_con_gruppo[anno_fine].append((1, mese_fine, gruppo))

                self.ultimo_regime = "spettacolo"
                self.ultimo_gruppo = gruppo

            elif giorni:
                # Record SENZA gruppo (es. Servizio Militare): raccogli periodi
                if anno_inizio == anno_fine:
                    periodi_senza_gruppo[anno_inizio].append((mese_inizio, mese_fine))
                    giorni_senza_gruppo_per_anno[anno_inizio] += giorni
                else:
                    # Periodo che attraversa anni: spezza per anno
                    periodi_senza_gruppo[anno_inizio].append((mese_inizio, 12))
                    periodi_senza_gruppo[anno_fine].append((1, mese_fine))
                    # Dividi giorni proporzionalmente (semplificato: tutto al primo anno)
                    giorni_senza_gruppo_per_anno[anno_inizio] += giorni

        # Calcola TEORICO per record CON gruppo (regole spettacolo)
        for anno, periodi in periodi_con_gruppo.items():
            # Unifica i mesi coperti (evita duplicati)
            mesi_coperti = set()
            gruppo_anno = None
            for mese_inizio, mese_fine, gruppo in periodi:
                for m in range(mese_inizio, mese_fine + 1):
                    mesi_coperti.add(m)
                gruppo_anno = gruppo

            # Unifica anche con eventuali periodi senza gruppo dello stesso anno
            if anno in periodi_senza_gruppo:
                for mese_inizio, mese_fine in periodi_senza_gruppo[anno]:
                    for m in range(mese_inizio, mese_fine + 1):
                        mesi_coperti.add(m)
                # Rimuovi da periodi_senza_gruppo perche' gia' conteggiato
                del periodi_senza_gruppo[anno]

            mesi = len(mesi_coperti)
            if mesi > 0 and gruppo_anno:
                giorni_teorici = self._calcola_teorico_spettacolo_con_mesi(anno, mesi_coperti, gruppo_anno)
                self.teorico_per_anno[anno] += giorni_teorici
                self.mesi_per_anno[anno] += mesi

        # Calcola TEORICO per record SENZA gruppo rimasti (non sovrapposti con gruppo)
        # Es. Servizio Militare: giorni reali = giorni teorici
        for anno, periodi in periodi_senza_gruppo.items():
            mesi_coperti = set()
            for mese_inizio, mese_fine in periodi:
                for m in range(mese_inizio, mese_fine + 1):
                    mesi_coperti.add(m)

            mesi = len(mesi_coperti)
            if mesi > 0:
                # Usa i giorni reali come teorici
                self.teorico_per_anno[anno] += giorni_senza_gruppo_per_anno[anno]
                self.mesi_per_anno[anno] += mesi

    def _calcola_teorico_spettacolo_con_mesi(self, anno, mesi_coperti, gruppo):
        """
        Calcola giorni teorici per Spettacolo considerando i mesi effettivi.
        Per il 1997 calcola proporzionalmente prima e dopo agosto.

        Regole:
        - Fino al 1992: Gruppo 1 = 60 gg/anno, Gruppo 2 = 180 gg/anno
        - 1993 - luglio 1997: Gruppo 1 = 120 gg/anno, Gruppo 2 = 260 gg/anno
        - Dal 1 agosto 1997 in poi:
          - Tempo determinato: Gruppo 1 = 120 gg/anno, Gruppo 2 = 260 gg/anno
          - Tempo indeterminato: sempre 312 gg/anno (indipendente dal gruppo)
        """
        if anno <= 1992:
            giorni_anno = 60 if gruppo == 1 else 180
            return round((giorni_anno / 12) * len(mesi_coperti))

        elif anno >= 1993 and anno < 1997:
            giorni_anno = 120 if gruppo == 1 else 260
            return round((giorni_anno / 12) * len(mesi_coperti))

        elif anno == 1997:
            # Anno di transizione: calcola proporzionalmente
            # Gen-Lug (mesi 1-7): regole 1993-1997 (sempre tempo determinato)
            # Ago-Dic (mesi 8-12): nuove regole (dipende da tempo ind/det)
            mesi_prima_agosto = [m for m in mesi_coperti if m <= 7]
            mesi_da_agosto = [m for m in mesi_coperti if m >= 8]

            # Prima di agosto: sempre regole vecchie
            giorni_prima = 120 if gruppo == 1 else 260
            totale_prima = giorni_prima * len(mesi_prima_agosto) / 12

            # Da agosto: dipende dal tipo contratto
            totale_dopo = 0
            for mese in mesi_da_agosto:
                if self._is_tempo_indeterminato(anno, mese):
                    totale_dopo += 312 / 12
                else:
                    giorni = 120 if gruppo == 1 else 260
                    totale_dopo += giorni / 12

            return round(totale_prima + totale_dopo)

        else:  # anno >= 1998
            # Calcola mese per mese per gestire passaggio a tempo indeterminato
            totale = 0
            for mese in mesi_coperti:
                if self._is_tempo_indeterminato(anno, mese):
                    totale += 312 / 12
                else:
                    giorni_anno = 120 if gruppo == 1 else 260
                    totale += giorni_anno / 12
            return round(totale)

    def _calcola_teorico_spettacolo(self, anno, mese_inizio, mesi, gruppo):
        """
        Versione per estensione anni futuri.
        Usa le regole post-1997 con gestione tempo indeterminato.
        """
        # Per anni futuri, calcola mese per mese
        totale = 0
        for i in range(mesi):
            mese = mese_inizio + i
            if self._is_tempo_indeterminato(anno, mese):
                totale += 312 / 12
            else:
                giorni_anno = 120 if gruppo == 1 else 260
                totale += giorni_anno / 12
        return round(totale)
    
    def _determina_range_anni(self):
        """Determina anno minimo e massimo dai dati"""
        tutti_anni = set(self.reale_per_anno.keys()) | set(self.teorico_per_anno.keys())
        if tutti_anni:
            self.anno_min = min(tutti_anni)
            self.anno_max = max(tutti_anni)

    def _estendi_a_obiettivo(self):
        """Estende il calcolo fino a raggiungere l'obiettivo contributivo (basato sul sesso)"""
        # Prima: completa l'ultimo anno lavorato a 12 mesi
        self._completa_ultimo_anno()

        # Calcola mesi gia' accumulati
        mesi_accumulati = sum(self.mesi_per_anno.values())

        if mesi_accumulati >= self.obiettivo_mesi:
            return  # Gia' raggiunto obiettivo

        mesi_mancanti = self.obiettivo_mesi - mesi_accumulati
        anno_corrente = self.anno_max + 1

        while mesi_mancanti > 0:
            mesi_anno = min(12, mesi_mancanti)
            self.mesi_per_anno[anno_corrente] = mesi_anno

            # Calcola giorni teorici in base all'ultimo regime
            if self.ultimo_regime == "generale":
                giorni_teorici = mesi_anno * 26
            else:  # spettacolo
                giorni_teorici = self._calcola_teorico_spettacolo(
                    anno_corrente, 1, mesi_anno, self.ultimo_gruppo or 2
                )

            self.teorico_per_anno[anno_corrente] = giorni_teorici
            self.reale_per_anno[anno_corrente] = 0  # Nessun contributo reale

            mesi_mancanti -= mesi_anno
            anno_corrente += 1

        self.anno_max = anno_corrente - 1

    def _completa_ultimo_anno(self):
        """Completa l'ultimo anno lavorato a 12 mesi nel TEORICO"""
        if not self.anno_max:
            return

        mesi_ultimo_anno = self.mesi_per_anno.get(self.anno_max, 0)
        if mesi_ultimo_anno > 0 and mesi_ultimo_anno < 12:
            # Completa a 12 mesi
            self.mesi_per_anno[self.anno_max] = 12

            # Ricalcola giorni teorici per l'anno completo (tutti i 12 mesi)
            if self.ultimo_regime == "generale":
                self.teorico_per_anno[self.anno_max] = 12 * 26  # 312
            else:  # spettacolo
                gruppo = self.ultimo_gruppo or 2
                mesi_completi = set(range(1, 13))  # tutti i 12 mesi
                self.teorico_per_anno[self.anno_max] = self._calcola_teorico_spettacolo_con_mesi(
                    self.anno_max, mesi_completi, gruppo
                )


# =============================================================================
# GENERAZIONE EXCEL
# =============================================================================

class GeneratoreExcel:
    """Classe per generare il file Excel con i risultati"""
    
    def __init__(self, risultati_calcolo):
        self.risultati = risultati_calcolo
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Contributi Previdenziali"
    
    def genera(self, output_path):
        """Genera il file Excel"""
        self._applica_stili()
        self._crea_headers()
        self._popola_dati()
        self._aggiungi_totali()
        self._imposta_larghezza_colonne()
        self.wb.save(output_path)
    
    def _applica_stili(self):
        """Definisce gli stili"""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill("solid", fgColor="4472C4")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center = Alignment(horizontal='center')
    
    def _crea_headers(self):
        """Crea le intestazioni"""
        headers = ["Anno", "Giorni REALI", "", "Anno", "Giorni TEORICI", "Mesi", "Anni e Mesi Cumulativi"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            if header:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center
                cell.border = self.border
    
    def _popola_dati(self):
        """Popola i dati per ogni anno"""
        anno_min = self.risultati["anno_min"]
        anno_max = self.risultati["anno_max"]

        if not anno_min or not anno_max:
            return

        mesi_cumulativi = 0

        for i, anno in enumerate(range(anno_min, anno_max + 1), 2):
            reale = self.risultati["reale"].get(anno, 0)
            teorico = self.risultati["teorico"].get(anno, 0)
            mesi = self.risultati["mesi"].get(anno, 0)

            # Colonne REALI (A, B)
            self.ws.cell(row=i, column=1, value=anno).border = self.border
            self.ws.cell(row=i, column=1).alignment = self.center
            self.ws.cell(row=i, column=2, value=reale).border = self.border
            self.ws.cell(row=i, column=2).alignment = self.center

            # Colonne TEORICI (D, E)
            self.ws.cell(row=i, column=4, value=anno).border = self.border
            self.ws.cell(row=i, column=4).alignment = self.center
            self.ws.cell(row=i, column=5, value=teorico).border = self.border
            self.ws.cell(row=i, column=5).alignment = self.center

            # Mesi (F)
            self.ws.cell(row=i, column=6, value=mesi).border = self.border
            self.ws.cell(row=i, column=6).alignment = self.center

            # Anni e Mesi Cumulativi (G)
            mesi_cumulativi += mesi
            anni_cum = mesi_cumulativi // 12
            mesi_cum = mesi_cumulativi % 12
            formato_cumulativo = f"{anni_cum}a {mesi_cum}m"
            self.ws.cell(row=i, column=7, value=formato_cumulativo).border = self.border
            self.ws.cell(row=i, column=7).alignment = self.center
    
    def _aggiungi_totali(self):
        """Aggiunge la riga dei totali"""
        anno_min = self.risultati["anno_min"]
        anno_max = self.risultati["anno_max"]

        if not anno_min or not anno_max:
            return

        num_anni = anno_max - anno_min + 1
        last_row = num_anni + 2

        # Totale REALE
        self.ws.cell(row=last_row, column=1, value="TOTALE").font = Font(bold=True)
        self.ws.cell(row=last_row, column=1).border = self.border
        self.ws.cell(row=last_row, column=2, value=f"=SUM(B2:B{last_row-1})").font = Font(bold=True)
        self.ws.cell(row=last_row, column=2).border = self.border
        self.ws.cell(row=last_row, column=2).alignment = self.center

        # Totale TEORICO
        self.ws.cell(row=last_row, column=4, value="TOTALE").font = Font(bold=True)
        self.ws.cell(row=last_row, column=4).border = self.border
        self.ws.cell(row=last_row, column=5, value=f"=SUM(E2:E{last_row-1})").font = Font(bold=True)
        self.ws.cell(row=last_row, column=5).border = self.border
        self.ws.cell(row=last_row, column=5).alignment = self.center

        # Totale Mesi
        self.ws.cell(row=last_row, column=6, value=f"=SUM(F2:F{last_row-1})").font = Font(bold=True)
        self.ws.cell(row=last_row, column=6).border = self.border
        self.ws.cell(row=last_row, column=6).alignment = self.center

        # Label finale per Anni e Mesi Cumulativi (usa obiettivo dal calcolo)
        obiettivo_label = self.risultati.get("obiettivo_label", "42a 10m")
        self.ws.cell(row=last_row, column=7, value=obiettivo_label).font = Font(bold=True)
        self.ws.cell(row=last_row, column=7).border = self.border
        self.ws.cell(row=last_row, column=7).alignment = self.center
    
    def _imposta_larghezza_colonne(self):
        """Imposta la larghezza delle colonne"""
        self.ws.column_dimensions['A'].width = 12
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 5
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 18
        self.ws.column_dimensions['F'].width = 10
        self.ws.column_dimensions['G'].width = 24


# =============================================================================
# MAIN
# =============================================================================

def main():
    # Parser argomenti
    parser = argparse.ArgumentParser(
        description="Estrae e calcola contributi previdenziali da PDF INPS",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempio:
    python contributi_inps.py certificazione.pdf                    # Tempo determinato
    python contributi_inps.py certificazione.pdf -ti                # Sempre tempo indeterminato
    python contributi_inps.py certificazione.pdf -ti 01/08/1997     # Tempo indeterminato dal 1/8/1997
        """
    )
    parser.add_argument("pdf", help="Percorso del file PDF INPS")
    parser.add_argument("-ti", "--tempo-indeterminato", nargs="?", const="sempre",
                        metavar="DD/MM/YYYY",
                        help="Tempo indeterminato: senza data = sempre, con data = da quella data")

    args = parser.parse_args()

    # Parsing tempo indeterminato
    tempo_indeterminato_da = None  # None = sempre tempo determinato
    if args.tempo_indeterminato:
        if args.tempo_indeterminato == "sempre":
            tempo_indeterminato_da = "sempre"
        else:
            # Valida formato data
            if not re.match(r'\d{2}/\d{2}/\d{4}', args.tempo_indeterminato):
                print(f"Errore: Formato data non valido: {args.tempo_indeterminato}")
                print("Usare formato DD/MM/YYYY")
                sys.exit(1)
            tempo_indeterminato_da = args.tempo_indeterminato

    # Verifica esistenza PDF
    if not os.path.exists(args.pdf):
        print(f"Errore: File non trovato: {args.pdf}")
        sys.exit(1)

    # Cartella output fissa
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    print("=" * 60)
    print("CALCOLO CONTRIBUTI PREVIDENZIALI INPS")
    print("=" * 60)

    # 1. Estrazione PDF
    print(f"\n[1/4] Estrazione dati da: {args.pdf}")
    estrattore = EstrattorePDF(args.pdf)
    dati = estrattore.estrai()

    # Decodifica sesso dal codice fiscale
    codice_fiscale = dati["metadata"].get("codice_fiscale")
    cognome = dati["metadata"].get("cognome")
    nome = dati["metadata"].get("nome")
    sesso = decodifica_sesso_da_cf(codice_fiscale)
    sesso_label = "Donna" if sesso == 'F' else "Uomo" if sesso == 'M' else "Non determinato"

    print(f"      - Codice Fiscale: {codice_fiscale}")
    if cognome and nome:
        print(f"      - Cognome e Nome: {cognome} {nome}")
    print(f"      - Sesso: {sesso_label}")
    if tempo_indeterminato_da:
        if tempo_indeterminato_da == "sempre":
            print(f"      - Contratto: Tempo indeterminato (sempre)")
        else:
            print(f"      - Contratto: Tempo indeterminato dal {tempo_indeterminato_da}")
    else:
        print(f"      - Contratto: Tempo determinato")
    print(f"      - Regime generale: {len(dati['regime_generale'])} record")
    print(f"      - Spettacolo: {len(dati['spettacolo'])} record")

    # Nome file basato su Cognome Nome, fallback a codice fiscale
    if cognome and nome:
        nome_file = f"{cognome} {nome}"
    elif codice_fiscale:
        nome_file = codice_fiscale
    else:
        # Fallback al nome del PDF se nulla trovato
        nome_file = os.path.splitext(os.path.basename(args.pdf))[0]

    json_path = os.path.join(output_dir, f"{nome_file}.json")
    excel_path = os.path.join(output_dir, f"{nome_file}.xlsx")

    # 2. Salvataggio JSON
    print(f"\n[2/4] Salvataggio JSON: {json_path}")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(dati, f, indent=2, ensure_ascii=False)

    # 3. Calcolo contributi (con sesso per determinare obiettivo)
    print(f"\n[3/4] Calcolo contributi...")
    calcolatore = CalcolatoreContributi(dati, sesso=sesso, tempo_indeterminato_da=tempo_indeterminato_da)
    risultati = calcolatore.calcola()

    # 4. Generazione Excel
    print(f"[4/4] Generazione Excel: {excel_path}")
    generatore = GeneratoreExcel(risultati)
    generatore.genera(excel_path)

    # Riepilogo
    totale_reale = sum(risultati["reale"].values())
    totale_teorico = sum(risultati["teorico"].values())
    totale_mesi = sum(risultati["mesi"].values())
    num_anni = risultati["anno_max"] - risultati["anno_min"] + 1 if risultati["anno_min"] else 0
    obiettivo_label = risultati.get("obiettivo_label", "42a 10m")

    print("\n" + "=" * 60)
    print("RIEPILOGO")
    print("=" * 60)
    if cognome and nome:
        print(f"Cognome e Nome:      {cognome} {nome}")
    print(f"Codice Fiscale:      {codice_fiscale}")
    print(f"Sesso:               {sesso_label}")
    print(f"Obiettivo:           {obiettivo_label}")
    print(f"Anni elaborati:      {num_anni} ({risultati['anno_min']} - {risultati['anno_max']})")
    print(f"Totale giorni REALI: {totale_reale}")
    print(f"Totale giorni TEORICI: {totale_teorico}")
    print(f"Totale mesi teorici: {totale_mesi} ({totale_mesi // 12}a {totale_mesi % 12}m)")
    print(f"\nFile generati:")
    print(f"  - {json_path}")
    print(f"  - {excel_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()